import streamlit as st
import pandas as pd
import requests
import time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

st.set_page_config(page_title="Lead Generation Tool", layout="wide")

# Custom CSS for professional styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1F4E78;
        margin-bottom: 1rem;
    }
    .subheader {
        font-size: 1.2rem;
        color: #444;
        margin-bottom: 1.5rem;
    }
    .success-box {
        background-color: #D4EDDA;
        border: 1px solid #C3E6CB;
        color: #155724;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    .error-box {
        background-color: #F8D7DA;
        border: 1px solid #F5C6CB;
        color: #721C24;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# App title
st.markdown('<div class="main-header">🎯 Lead Generation Tool</div>', unsafe_allow_html=True)
st.markdown('<div class="subheader">Find businesses and their executives with ease</div>', unsafe_allow_html=True)

class LeadGeneratorAPI:
    def __init__(self, google_key, serpapi_key):
        self.google_key = google_key
        self.serpapi_key = serpapi_key

    def search_google_places(self, query, max_results=20):
        """Search for businesses using Google Places API"""
        businesses = []

        try:
            url = "https://places.googleapis.com/v1/places:searchText"
            headers = {"Content-Type": "application/json", "X-Goog-Api-Key": self.google_key}

            response = requests.post(
                url,
                json={"textQuery": query},
                headers=headers,
                timeout=20
            )

            if response.status_code == 200:
                data = response.json()
                for place in data.get("places", [])[:max_results]:
                    business = {
                        "Business Name": place.get("displayName", {}).get("text", "N/A"),
                        "Address": place.get("formattedAddress", "N/A"),
                        "Phone": place.get("internationalPhoneNumber", "N/A"),
                        "Website": place.get("websiteUri", "N/A"),
                        "About": place.get("editorialSummary", {}).get("text", "")[:300] if place.get("editorialSummary") else "N/A"
                    }
                    businesses.append(business)

        except Exception as e:
            logger.error(f"Google Places error: {str(e)}")
            st.warning(f"⚠️ Error searching Google Places: {str(e)}")

        return businesses

    def search_linkedin_executive(self, company_name):
        """Search for executives on LinkedIn via SerpAPI with multiple strategies"""
        search_strategies = [
            f"{company_name} CEO linkedin",
            f"{company_name} founder linkedin",
            f"{company_name} director linkedin",
            f"site:linkedin.com {company_name} CEO",
            f"site:linkedin.com {company_name} founder",
            f"site:linkedin.com/in {company_name}",
            f"{company_name} executive linkedin",
        ]

        for query in search_strategies:
            for attempt in range(2):
                try:
                    params = {
                        "q": query,
                        "engine": "google",
                        "api_key": self.serpapi_key,
                        "num": 10
                    }

                    response = requests.get(
                        "https://api.serpapi.com/search",
                        params=params,
                        timeout=20
                    )

                    if response.status_code == 200:
                        data = response.json()

                        # Check organic results
                        for result in data.get("organic_results", []):
                            link = result.get("link", "").lower()
                            title = result.get("title", "")

                            if "linkedin.com" in link and ("/in/" in link or "/company/" in link):
                                name = title.split("|")[0].strip() if "|" in title else title
                                return {
                                    "Executive Name": name[:80],
                                    "Role": "Professional",
                                    "LinkedIn Profile": result.get("link", ""),
                                    "About": result.get("snippet", "")[:300]
                                }

                        # Check people results
                        for person in data.get("people_results", []):
                            name = person.get("name", "")
                            link = person.get("link", "")
                            if name and link and "linkedin" in link.lower():
                                return {
                                    "Executive Name": name[:80],
                                    "Role": person.get("title", "Professional"),
                                    "LinkedIn Profile": link,
                                    "About": person.get("snippet", "")[:300]
                                }

                    if attempt == 0:
                        time.sleep(1.5)

                except Exception as e:
                    logger.error(f"SerpAPI error for query '{query}': {str(e)}")
                    if attempt == 0:
                        time.sleep(1.5)

        return {
            "Executive Name": "Not Found",
            "Role": "Not Found",
            "LinkedIn Profile": "Not Found",
            "About": ""
        }

    def generate_leads(self, business_type, location, num_results=20):
        """Generate complete lead list with businesses and executives"""
        query = f"{business_type} in {location}"

        # Search for businesses
        st.info(f"🔍 Searching for {num_results} businesses matching '{query}'...")
        businesses = self.search_google_places(query, num_results)

        if not businesses:
            st.error(f"❌ No businesses found for '{query}'")
            return None, None

        st.success(f"✅ Found {len(businesses)} businesses!")

        # Search for executives
        st.info("🔎 Searching for executives at these companies...")
        progress_bar = st.progress(0)

        executives = []
        for idx, business in enumerate(businesses):
            company_name = business.get("Business Name", "")
            executive = self.search_linkedin_executive(company_name)
            executive["Target_Company_Name"] = company_name
            executives.append(executive)
            progress_bar.progress((idx + 1) / len(businesses))
            time.sleep(1)

        st.success(f"✅ Completed executive search!")

        return businesses, executives


def create_professional_excel(businesses_df, executives_df):
    """Create a professionally formatted Excel workbook with two sheets"""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Business Leads sheet
    ws_business = wb.create_sheet("Business Leads")
    headers_business = ["Business Name", "Website", "Phone", "Address", "About"]
    ws_business.append(headers_business)

    for _, row in businesses_df.iterrows():
        ws_business.append([
            row.get("Business Name", ""),
            row.get("Website", ""),
            row.get("Phone", ""),
            row.get("Address", ""),
            row.get("About", "")
        ])

    # Executive Leads sheet
    ws_exec = wb.create_sheet("Executive Leads")
    headers_exec = ["Person Name", "Target Company Name", "Role", "LinkedIn Profile", "About"]
    ws_exec.append(headers_exec)

    for _, row in executives_df.iterrows():
        ws_exec.append([
            row.get("Executive Name", ""),
            row.get("Target_Company_Name", ""),
            row.get("Role", ""),
            row.get("LinkedIn Profile", ""),
            row.get("About", "")
        ])

    # Format both sheets
    for ws in [ws_business, ws_exec]:
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                if row_idx % 2 == 0:
                    cell.fill = light_gray_fill

        column_widths = {"A": 25, "B": 30, "C": 20, "D": 35, "E": 30}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# Initialize session state
if "api_initialized" not in st.session_state:
    st.session_state.api_initialized = False
    st.session_state.businesses = None
    st.session_state.executives = None

# Check for API credentials
try:
    google_key = st.secrets["google_places_api_key"]
    serpapi_key = st.secrets["serpapi_api_key"]
    api = LeadGeneratorAPI(google_key, serpapi_key)
    st.session_state.api_initialized = True
except KeyError as e:
    st.markdown('<div class="error-box">❌ API credentials not found in Streamlit Secrets</div>', unsafe_allow_html=True)
    st.error(f"Missing credential: {str(e)}")
    st.info("""
    **To fix this:**
    1. Go to your Streamlit Cloud app settings → Secrets
    2. Add these lines in TOML format:
    ```toml
    google_places_api_key = "YOUR_GOOGLE_KEY"
    serpapi_api_key = "YOUR_SERPAPI_KEY"
    ```
    """)
    st.stop()

# Main interface
col1, col2, col3 = st.columns(3)

with col1:
    business_type = st.text_input("Business Type", placeholder="e.g., mental health recovery centers", value="")

with col2:
    location = st.text_input("Location", placeholder="e.g., Texas, California", value="")

with col3:
    num_results = st.number_input("Number of Results", min_value=5, max_value=50, value=10, step=5)

# Search button
if st.button("🔍 Generate Leads", use_container_width=True):
    if not business_type or not location:
        st.error("Please enter both Business Type and Location")
    else:
        with st.spinner("Processing your request..."):
            businesses, executives = api.generate_leads(business_type, location, num_results)

            if businesses and executives:
                st.session_state.businesses = businesses
                st.session_state.executives = executives

                st.markdown("### 📊 Results Summary")
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Businesses Found", len(businesses))
                with col2:
                    executives_found = sum(1 for e in executives if e["Executive Name"] != "Not Found")
                    st.metric("Executives Found", executives_found)

                st.subheader("📍 Business Leads")
                business_df = pd.DataFrame(businesses)
                st.dataframe(business_df, use_container_width=True, hide_index=True)

                st.subheader("👤 Executive Leads")
                exec_df = pd.DataFrame(executives)
                st.dataframe(exec_df, use_container_width=True, hide_index=True)

# Download section
if st.session_state.businesses and st.session_state.executives:
    st.markdown("---")
    st.subheader("📥 Download Results")

    businesses_df = pd.DataFrame(st.session_state.businesses)
    executives_df = pd.DataFrame(st.session_state.executives)

    excel_file = create_professional_excel(businesses_df, executives_df)
    filename = f"{business_type} {location} leads.xlsx".replace(" ", "_").lower()

    st.download_button(
        label="📊 Download Excel File",
        data=excel_file,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.markdown('<div class="success-box">✅ Ready to download! Click the button above to get your lead list in Excel format.</div>', unsafe_allow_html=True)
